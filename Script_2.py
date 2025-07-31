# ==================================================================================================
# Script_2.py
#
# DESCRIPTION:
# This script functions as a web scraper for the sofifa.com website.
# Its main focus is to collect player lists from a team page, then
# visit each player's page to gather specific data: Name, Value,
# and Wage.
# The collected data will be saved in three separate formats: Text (.txt), JSON (.json),
# and Excel (.xlsx) using the openpyxl library.
#
# USAGE:
# python Script_2.py --url [TEAM_URL_ON_SOFIFA] --output [OUTPUT_FOLDER_NAME]
#
# EXAMPLE:
# python Script_2.py --url "https://sofifa.com/team/10/manchester-city" --output "C:\Users\proye\Downloads\COBA REMOVE BG"
# ==================================================================================================


# --------------------------------------------------------------------------------------------------
# SECTION 1: LIBRARY IMPORTS
# Importing all necessary libraries to run the script.
# --------------------------------------------------------------------------------------------------
import os                     # Used for system operations, such as creating output directories.
import time                   # Used to introduce delays between requests
                              # to avoid overloading servers and reduce bot detection risk.
import json                   # Used for working with JSON data format,
                              # particularly when saving scraping results.
import random                 # Used for random selection,
                              # such as User-Agents to disguise requests.
import argparse               # Used for parsing command-line arguments
                              # (e.g., team URL and output directory).
import openpyxl               # Main library for reading and writing Excel (.xlsx) files.
from openpyxl.utils import get_column_letter # Utility from openpyxl to convert column indexes
                                             # to Excel column letters (e.g., 1 -> A, 2 -> B).
from bs4 import BeautifulSoup # Library for parsing HTML and XML documents. Used to
                              # extract data from web page sources.
import undetected_chromedriver as uc # Modified version of Selenium ChromeDriver designed
                                     # to avoid detection as a bot by websites.
from selenium.webdriver.common.by import By # Used to specify element search strategies
                                            # on web pages (e.g., by CSS Selector, ID, etc.).
from selenium.webdriver.support.ui import WebDriverWait # Used to wait for specific conditions
                                                        # before continuing execution (e.g., waiting for elements to appear).
from selenium.webdriver.support import expected_conditions as EC # Collection of conditions WebDriverWait can check,
                                                                 # such as element presence or visibility.
import sys                    # Used to access system-specific parameters and functions,
                              # such as console output configuration.
import io                     # Used for managing I/O streams,
                              # particularly for setting console output encoding.
import re                     # Used for regular expression operations,
                              # e.g., extracting player IDs from URLs.


# --------------------------------------------------------------------------------------------------
# SECTION 2: GLOBAL CONFIGURATION AND HELPER FUNCTIONS
# Variables and small functions used throughout the script for utility purposes.
# --------------------------------------------------------------------------------------------------

# Fix encoding issues for console output to correctly display special characters (e.g., currency symbols).
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def log(msg):
    """
    Simple function to print messages to the console.
    Using `flush=True` ensures messages are displayed immediately without buffering,
    useful for real-time progress tracking.

    Args:
        msg (str): Message to print to the console.
    """
    print(msg, flush=True)

# List of different User-Agents to disguise HTTP requests.
# Randomly selecting a User-Agent helps avoid detection as a bot.
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.5938.62 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15"
]


# --------------------------------------------------------------------------------------------------
# SECTION 3: MAIN FUNCTIONS
# Collection of functions that form the core logic of this web scraper.
# --------------------------------------------------------------------------------------------------
def setup_driver():
    """
    Initializes and configures Chrome driver using `undetected_chromedriver`.
    Sets various options to ensure optimal driver operation, especially in headless mode
    (without visible browser interface).

    Returns:
        uc.Chrome: Configured Selenium driver object ready for use.
    """
    options = uc.ChromeOptions()
    options.add_argument("--no-sandbox")            # Disable sandbox mode for certain environments (e.g., Docker).
    options.add_argument("--disable-dev-shm-usage") # Fix shared memory issues in Linux environments.
    options.add_argument("--disable-gpu")           # Disable GPU acceleration (useful for headless/servers).
    options.add_argument("--window-size=1920,1080") # Set browser window size.
    options.add_argument("--disable-extensions")    # Disable browser extensions.
    options.add_argument("--disable-infobars")      # Hide the "Chrome is being controlled by automated test software" info bar.
    options.add_argument("--start-maximized")       # Start browser in maximized mode.
    options.add_argument(f"--user-agent={random.choice(USER_AGENTS)}") # Set random User-Agent.
    options.add_argument("--headless=new")          # Run browser in headless mode (without GUI).
    driver = uc.Chrome(options=options)
    return driver

def scroll_to_bottom(driver):
    """
    Scrolls the web page from top to bottom slowly until no new content
    is loaded. This is important for sites using lazy loading, where content only
    appears when the user scrolls down.

    Args:
        driver: Active Selenium driver object.
    """
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        # Scroll to the bottom of the page
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # Introduce random delay to allow content to load
        time.sleep(random.uniform(2, 3))
        # Get new page height
        new_height = driver.execute_script("return document.body.scrollHeight")
        # If page height hasn't changed, we've reached the bottom
        if new_height == last_height:
            break
        last_height = new_height

def get_player_urls_from_team(team_url, driver, max_retries=3):
    """
    Retrieves all unique URLs of player profiles listed on a sofifa.com team page.
    This function attempts multiple times if failures occur.

    Args:
        team_url (str): Complete URL of the team page on sofifa.com.
        driver: Selenium driver object used for web navigation.
        max_retries (int): Maximum number of attempts to fetch data if errors occur.

    Returns:
        list: List of unique player profile URLs. Returns empty list if all attempts fail.
    """
    attempt = 0
    while attempt < max_retries:
        try:
            time.sleep(random.uniform(4, 7)) # Random delay before loading page
            log(f"\n🔍 Loading team page: {team_url}")
            driver.get(team_url)

            scroll_to_bottom(driver) # Scroll to bottom to load all players

            # Save page HTML source for debugging (this file will be created in the same directory)
            with open("page_dump.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)

            # Wait until <tbody> elements (table body) appear on the page, indicating main content is loaded.
            WebDriverWait(driver, 20).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tbody"))
            )

            # Find all <a> elements (links) pointing to player pages.
            # Selector `td.col-name a[href^='/player/']` and `tbody a[href^='/player/']`
            # ensures we only get links to player profiles.
            players = driver.find_elements(By.CSS_SELECTOR, "td.col-name a[href^='/player/'], tbody a[href^='/player/']")
            # Get 'href' attribute from each link and remove additional URL parameters
            # (e.g., "?type=all") to get clean player profile URLs.
            player_urls = [a.get_attribute("href").split("?")[0] for a in players if "/player/" in a.get_attribute("href")]

            log(f"🔗 Number of player links found: {len(player_urls)}")
            for link in player_urls:
                log("🔗 " + link)

            if not player_urls:
                log("⚠️ No player elements found on page. Ensure selector is correct or page has loaded.")

            # Using `dict.fromkeys` then converting back to list
            # is an efficient way to remove duplicates while maintaining order.
            return list(dict.fromkeys(player_urls))

        except Exception as e:
            attempt += 1
            log(f"❌ Failed to retrieve player list from {team_url} (Attempt {attempt}/{max_retries}): {e}")
            time.sleep(3) # Delay before retrying

    return [] # Return empty list if all attempts fail

def scrape_name_value_wage(player_url, driver, max_retries=3):
    """
    Retrieves Name, Value, and Wage from a player's profile page.
    This function also attempts multiple times if failures occur.

    Args:
        player_url (str): Complete URL of player profile page on sofifa.com.
        driver: Selenium driver object used for web navigation.
        max_retries (int): Maximum number of attempts to fetch data if errors occur.

    Returns:
        tuple: Returns tuple containing (player_id, name, value, wage).
               Returns default values ("-") if data not found or errors occur.
    """
    attempt = 0
    player_id = "-"
    # Extract player ID from URL using regular expression.
    match = re.search(r"/player/(\d+)", player_url)
    if match:
        player_id = match.group(1) # Player ID found in first regex group

    while attempt < max_retries:
        try:
            time.sleep(random.uniform(4, 7)) # Random delay before loading player page
            driver.get(player_url)
            log(f"🔍 Fetching data from player: {player_url}")

            # Wait until element with CSS Selector 'div.grid' appears,
            # indicating the main section of player page has loaded.
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div.grid'))
            )

            # Use BeautifulSoup to parse HTML from the page loaded by driver.
            soup = BeautifulSoup(driver.page_source, "html.parser")

            # Find <h1> element which typically contains player name.
            name_tag = soup.select_one("h1")
            name = name_tag.text.strip() if name_tag else "-" # Get name text or '-' if not found
            value = wage = "-" # Initialize value and wage

            # Iterate through each column in div.grid to find "Value" and "Wage".
            for col in soup.select("div.grid div.col"):
                sub = col.find("div", class_="sub") # Find element with class "sub" (label)
                if sub:
                    label = sub.text.strip() # Get label text (e.g., "Value", "Wage")
                    value_el = col.find("em") # Find <em> element containing value or wage
                    if value_el:
                        if label == "Value":
                            value = value_el.text.strip()
                        elif label == "Wage":
                            wage = value_el.text.strip()

            return player_id, name, value, wage

        except Exception as e:
            attempt += 1
            log(f"⚠️ Failed to fetch data from {player_url} (Attempt {attempt}/{max_retries}): {e}")
            time.sleep(3) # Delay before retrying

    # If all attempts fail, return default values
    return player_id, "-", "-", "-"


# --------------------------------------------------------------------------------------------------
# SECTION 4: MAIN EXECUTION BLOCK
# This section runs when the script is executed directly from the command line.
# It is the main workflow of the script.
# --------------------------------------------------------------------------------------------------
if __name__ == "__main__":
    # Create argument parser to accept command-line inputs.
    parser = argparse.ArgumentParser(description="Scraper for player Value & Wage data from sofifa.com")
    # Required argument: Sofifa team URL.
    parser.add_argument("--url", required=True, help="Sofifa team URL link")
    # Optional argument: output folder, default is "output".
    parser.add_argument("--output", default="output", help="Output data folder")
    # Parse arguments provided by user.
    args = parser.parse_args()

    team_url = args.url      # Team URL to scrape.
    output_dir = args.output # Directory to save results.

    log(f"\n🔗 Starting scraping from: {team_url}")
    start_time = time.time() # Record script execution start time.

    # Initialize browser driver.
    driver = setup_driver()
    # Retrieve list of player URLs from team page.
    player_urls = get_player_urls_from_team(team_url, driver)

    # Check if any player URLs were found. If not, exit program.
    if not player_urls:
        log("❌ No players found. Program terminating.")
        driver.quit() # Close browser driver.
        exit()        # Exit script.

    log(f"✅ Found {len(player_urls)} unique players. Starting scraping process.")
    results = []          # List to store all successfully scraped player data.
    total_players = len(player_urls) # Total number of players to scrape.

    # Iterate through each found player URL to fetch their data.
    for i, url in enumerate(player_urls, start=1):
        t0 = time.time() # Start time for scraping this player.
        log(f"\n➡️ Player {i}/{total_players}")
        # Call function to scrape player name, value, and wage.
        player_id, name, value, wage = scrape_name_value_wage(url, driver)
        log(f"📝 {player_id} | {name} | Value: {value} | Wage: {wage}")
        # Add player data to results list.
        results.append({"ID": player_id, "Name": name, "Value": value, "Wage": wage})

        t1 = time.time() # Completion time for scraping this player.
        # Calculate average time per player and estimated completion time (ETA).
        avg_time_per_player = (t1 - start_time) / i
        eta = avg_time_per_player * (total_players - i)
        log(f"⏳ ETA: {eta:.1f} seconds")

        time.sleep(random.uniform(1.0, 2.5)) # Random delay between player data fetches.

    driver.quit() # Close browser driver after all scraping processes complete.

    # If any data was successfully scraped, save to files.
    if results:
        # Create output directory if it doesn't exist.
        os.makedirs(output_dir, exist_ok=True)
        # Extract team name from URL to use as prefix for output filenames.
        team_name = team_url.strip("/").split("/")[-1]
        filename_prefix = f"SCRIPT_2_{team_name}"

        # --- Save to Text File (.txt) ---
        output_txt = os.path.join(output_dir, f"{filename_prefix}.txt")
        with open(output_txt, "w", encoding="utf-8") as f:
            for player in results:
                f.write(f"{player['ID']} | {player['Name']} | Value: {player['Value']} | Wage: {player['Wage']}\n")
        log(f"💾 Data successfully saved to: {output_txt}")


        # --- Save to JSON File (.json) ---
        output_json = os.path.join(output_dir, f"{filename_prefix}.json")
        with open(output_json, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        log(f"💾 Data successfully saved to: {output_json}")


        # --- Save to Excel File (.xlsx) ---
        output_xlsx = os.path.join(output_dir, f"{filename_prefix}.xlsx")
        wb = openpyxl.Workbook() # Create new Excel workbook.
        ws = wb.active           # Get active worksheet.
        ws.title = "Players"     # Set worksheet title.
        ws.append(["ID", "Name", "Value", "Wage"]) # Add column headers.
        for row in results:
            ws.append([row["ID"], row["Name"], row["Value"], row["Wage"]]) # Add each row of data.

        # Set column widths to fit longest content in each column.
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2 # Add 2 for padding.
        wb.save(output_xlsx) # Save Excel workbook.
        log(f"💾 Data successfully saved to: {output_xlsx}")


        total_time = time.time() - start_time # Calculate total execution time.
        log(f"\n✅ Scraping process completed in {total_time:.1f} seconds.")
        log(f"📁 Output files available in '{output_dir}':")
        log(f" - {filename_prefix}.txt")
        log(f" - {filename_prefix}.json")
        log(f" - {filename_prefix}.xlsx")
    else:
        log("⚠️ No player data was saved because no data was scraped.")